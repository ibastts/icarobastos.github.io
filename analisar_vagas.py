"""
Analisador de Compatibilidade de Vagas
----------------------------------------
Le uma lista de vagas (coletadas manualmente ou exportadas de sites de
emprego) e calcula um score de compatibilidade com o seu perfil, gerando
uma planilha Excel organizada e ordenada por compatibilidade.

COMO USAR
1. Preencha o arquivo "vagas_input.xlsx" (veja o modelo gerado junto com
   este script) com as vagas que você encontrar. Colunas esperadas:
     Vaga | Empresa | Local | Modalidade | Salario | Beneficios | Descricao | Link
2. Rode: python analisar_vagas.py
3. O resultado sai em "vagas_compatibilidade.xlsx", ordenado da vaga mais
   compatível para a menos compatível.

Você pode (e deve) ajustar as listas de palavras-chave abaixo conforme for
percebendo o que realmente pesa nas vagas que te interessam.
"""

import re
import unicodedata

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ARQUIVO_ENTRADA = "vagas_input.xlsx"
ARQUIVO_SAIDA = "vagas_compatibilidade.xlsx"

# ---------------------------------------------------------------------------
# PALAVRAS-CHAVE — ajuste livremente conforme seu perfil e o que for buscando
# ---------------------------------------------------------------------------

# Termos ligados a Departamento Pessoal / Administrativo
KEYWORDS_DP = [
    "departamento pessoal", "dp", "folha de pagamento", "admissao",
    "rescisao", "ferias", "ponto eletronico", "beneficios", "esocial",
    "e-social", "rotinas trabalhistas", "clt", "encargos", "protheus",
    "recursos humanos", "rh",
]

# Termos ligados a Dados (área de maior interesse real)
KEYWORDS_DADOS = [
    "dados", "excel avancado", "tabela dinamica", "power bi", "powerbi",
    "sql", "python", "analise de dados", "dashboard", "kpi", "bi",
    "business intelligence", "planilhas", "relatorios gerenciais",
]

# Termos ligados a Fiscal / Financeiro / Faturamento (area de maior encaixe hoje)
KEYWORDS_FISCAL_FINANCEIRO = [
    "fiscal", "faturamento", "nota fiscal", "notas fiscais", "e-cac", "ecac",
    "e115", "cce", "carta de correcao", "conciliacao bancaria",
    "controle financeiro", "centro de custo", "contas a pagar",
    "contas a receber", "tesouraria", "tributario", "impostos", "nfe",
    "simples nacional", "obrigacoes acessorias",
]

# Termos ligados a habilidades administrativas gerais
KEYWORDS_ADMIN = [
    "excel", "pacote office", "word", "powerpoint", "contratos",
    "organizacao", "rotinas administrativas", "cadastro", "documentos",
    "comercial", "atendimento", "negocios", "vendas internas",
]

# Termos que indicam modalidade remota (bonus extra no score)
KEYWORDS_REMOTO = ["home office", "remoto", "home-office", "trabalho remoto"]

# Pesos de cada grupo no score final (soma = 100)
PESO_DP = 15
PESO_FISCAL_FINANCEIRO = 35
PESO_DADOS = 30
PESO_ADMIN = 15
PESO_REMOTO = 5


def normalizar(texto: str) -> str:
    """Remove acentos e baixa a caixa, para comparar palavras-chave sem
    depender de acentuação exata no texto da vaga."""
    if not isinstance(texto, str):
        return ""
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto.lower()


def contar_ocorrencias(texto_normalizado: str, termos: list[str]) -> int:
    encontrados = 0
    for termo in termos:
        termo_norm = normalizar(termo)
        if re.search(r"\b" + re.escape(termo_norm) + r"\b", texto_normalizado):
            encontrados += 1
    return encontrados


# Cidades onde presencial funciona (só a sua)
CIDADES_PRESENCIAL = ["ilheus"]

# Cidades onde hibrido funciona (raio de deslocamento viavel)
CIDADES_HIBRIDO = ["ilheus", "itabuna"]


def checar_elegibilidade(modalidade: str, local: str) -> str:
    """Verifica se a vaga e' viavel dado o tipo de contratacao e a cidade."""
    modalidade_norm = normalizar(modalidade)
    local_norm = normalizar(local)

    eh_remoto = contar_ocorrencias(modalidade_norm, KEYWORDS_REMOTO) > 0
    eh_hibrido = "hibrido" in modalidade_norm
    eh_presencial = "presencial" in modalidade_norm

    if eh_remoto:
        return "Apta"
    if eh_hibrido:
        return "Apta" if any(c in local_norm for c in CIDADES_HIBRIDO) else "Fora da area (hibrido)"
    if eh_presencial:
        return "Apta" if any(c in local_norm for c in CIDADES_PRESENCIAL) else "Fora da area (presencial)"

    # Modalidade nao informada com clareza: nao bloqueia, mas sinaliza pra conferir
    return "Verificar modalidade"


def calcular_score(vaga: str, descricao: str, modalidade: str, local: str) -> tuple[int, str]:
    texto_completo = normalizar(f"{vaga} {descricao}")
    texto_modalidade = normalizar(f"{modalidade} {local}")

    score_dp = min(contar_ocorrencias(texto_completo, KEYWORDS_DP) / 4, 1) * PESO_DP
    score_fiscal_financeiro = min(contar_ocorrencias(texto_completo, KEYWORDS_FISCAL_FINANCEIRO) / 4, 1) * PESO_FISCAL_FINANCEIRO
    score_dados = min(contar_ocorrencias(texto_completo, KEYWORDS_DADOS) / 4, 1) * PESO_DADOS
    score_admin = min(contar_ocorrencias(texto_completo, KEYWORDS_ADMIN) / 4, 1) * PESO_ADMIN
    score_remoto = PESO_REMOTO if contar_ocorrencias(texto_modalidade, KEYWORDS_REMOTO) > 0 else 0

    score_total = round(score_dp + score_fiscal_financeiro + score_dados + score_admin + score_remoto)

    elegibilidade = checar_elegibilidade(modalidade, local)
    if elegibilidade.startswith("Fora da area"):
        score_total = 0

    if score_total >= 70:
        classificacao = "Alta"
    elif score_total >= 40:
        classificacao = "Media"
    else:
        classificacao = "Baixa"

    return score_total, classificacao, elegibilidade


def gerar_modelo_entrada():
    """Cria um arquivo de exemplo caso o usuário ainda não tenha um."""
    exemplo = pd.DataFrame([
        {
            "Vaga": "Assistente de Departamento Pessoal",
            "Empresa": "Empresa Exemplo",
            "Local": "Remoto",
            "Modalidade": "Home office",
            "Salario": "R$ 2.200,00",
            "Beneficios": "Vale refeicao, vale transporte, plano de saude",
            "Descricao": "Rotinas de admissao, ferias, rescisao, folha de pagamento e controle de ponto. Excel intermediario.",
            "Link": "https://exemplo.com/vaga",
        }
    ])
    exemplo.to_excel(ARQUIVO_ENTRADA, index=False)
    print(f"Nao encontrei '{ARQUIVO_ENTRADA}'. Criei um modelo de exemplo — preencha com suas vagas e rode o script de novo.")


def main():
    try:
        df = pd.read_excel(ARQUIVO_ENTRADA)
    except FileNotFoundError:
        gerar_modelo_entrada()
        return

    colunas_esperadas = ["Vaga", "Empresa", "Local", "Modalidade", "Salario", "Beneficios", "Descricao", "Link"]
    for col in colunas_esperadas:
        if col not in df.columns:
            df[col] = ""

    resultados = df.apply(
        lambda row: calcular_score(row["Vaga"], row["Descricao"], row["Modalidade"], row["Local"]),
        axis=1,
    )
    df["Compatibilidade (%)"] = [r[0] for r in resultados]
    df["Classificacao"] = [r[1] for r in resultados]
    df["Elegibilidade"] = [r[2] for r in resultados]

    df = df.sort_values("Compatibilidade (%)", ascending=False).reset_index(drop=True)
    df = df[["Vaga", "Empresa", "Local", "Modalidade", "Salario", "Beneficios", "Compatibilidade (%)", "Classificacao", "Elegibilidade", "Descricao", "Link"]]

    df.to_excel(ARQUIVO_SAIDA, index=False, sheet_name="Vagas")
    formatar_planilha(ARQUIVO_SAIDA)
    print(f"Pronto! {len(df)} vaga(s) analisada(s). Resultado em '{ARQUIVO_SAIDA}'.")


def formatar_planilha(caminho: str):
    wb = load_workbook(caminho)
    ws = wb["Vagas"]

    fonte_padrao = "Arial"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name=fonte_padrao, bold=True, color="FFFFFF")

    fill_alta = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_media = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_baixa = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    class_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Classificacao":
            class_col_idx = idx
        cell_font = Font(name=fonte_padrao)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(name=fonte_padrao)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if class_col_idx:
            valor = row[class_col_idx - 1].value
            fill = {"Alta": fill_alta, "Media": fill_media, "Baixa": fill_baixa}.get(valor)
            if fill:
                for cell in row:
                    cell.fill = fill

    larguras = {"A": 32, "B": 22, "C": 16, "D": 16, "E": 16, "F": 30, "G": 16, "H": 14, "I": 22, "J": 55, "K": 30}
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    ws.freeze_panes = "A2"
    wb.save(caminho)


if __name__ == "__main__":
    main()
