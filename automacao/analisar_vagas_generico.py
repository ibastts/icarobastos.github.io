"""
Analisador de Compatibilidade de Vagas
----------------------------------------
Le uma lista de vagas (coletadas manualmente ou exportadas de sites de
emprego) e calcula um score de compatibilidade com o seu perfil, gerando
uma planilha Excel organizada e ordenada por compatibilidade.

IMPORTANTE: os valores de cidade e as areas/palavras-chave configurados
abaixo sao apenas um EXEMPLO ILUSTRATIVO (calibrados com a realidade de
quem criou este script — Ilheus/Itabuna, BA, e um perfil administrativo/
fiscal/dados). Antes de usar, edite o bloco "CONFIGURACAO" com a SUA
propria cidade e as SUAS proprias areas de interesse.

COMO USAR
1. Edite o bloco "CONFIGURACAO" logo abaixo com a sua cidade e as suas
   areas/habilidades. E' a UNICA parte do arquivo que voce precisa mexer.
2. Preencha o arquivo "vagas_input.xlsx" (um modelo e' criado automaticamente
   na primeira vez que voce rodar o script) com as vagas que encontrar.
   Colunas esperadas:
     Vaga | Empresa | Local | Modalidade | Salario | Beneficios | Descricao | Link
3. Rode: python analisar_vagas.py
4. O resultado sai em "vagas_compatibilidade.xlsx", ordenado da vaga mais
   compativel para a menos compativel.
"""

import re
import unicodedata

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# =============================================================================
# CONFIGURACAO — EDITE SOMENTE ESTE BLOCO COM OS SEUS DADOS
# =============================================================================

# 1) SUA CIDADE
# ATENCAO: os valores abaixo (Ilheus/Itabuna) sao apenas um EXEMPLO
# ILUSTRATIVO, calibrado com a realidade de quem criou este script. Antes
# de usar, TROQUE pela sua propria cidade — sem isso, o calculo de
# elegibilidade vai considerar a localizacao errada para o seu caso.
#
# MINHA_CIDADE_PRESENCIAL: cidade(s) onde voce aceita vaga 100% presencial
# MINHA_CIDADE_HIBRIDO: cidade(s) onde voce aceita vaga hibrida (normalmente
#   um raio um pouco maior do que o presencial, incluindo cidades vizinhas)
MINHA_CIDADE_PRESENCIAL = ["ilheus"]  # EXEMPLO — troque pela sua cidade
MINHA_CIDADE_HIBRIDO = ["ilheus", "itabuna"]  # EXEMPLO — troque pelas suas cidades

# 2) SUAS AREAS / HABILIDADES
# ATENCAO: os grupos e palavras-chave abaixo tambem sao apenas um EXEMPLO
# ILUSTRATIVO (perfil administrativo/fiscal/dados). Cada grupo representa
# uma area de interesse ou conjunto de habilidades. Para cada grupo, defina:
#   - um peso (quanto essa area importa no score final — a soma de todos os
#     pesos, incluindo o de modalidade remota, deve dar 100)
#   - uma lista de palavras-chave que costumam aparecer em vagas dessa area
#
# Adicione, remova ou renomeie grupos livremente. Exemplo pronto abaixo
# (perfil administrativo/fiscal/dados) — substitua pelo seu:
GRUPOS_DE_INTERESSE = {
    "Area 1": {
        "peso": 35,
        "palavras_chave": [
            "fiscal", "faturamento", "nota fiscal", "notas fiscais",
            "conciliacao bancaria", "controle financeiro", "contas a pagar",
            "contas a receber", "tesouraria", "tributario", "impostos",
        ],
    },
    "Area 2": {
        "peso": 30,
        "palavras_chave": [
            "dados", "excel avancado", "tabela dinamica", "power bi",
            "sql", "python", "analise de dados", "dashboard", "kpi",
            "business intelligence",
        ],
    },
    "Area 3": {
        "peso": 15,
        "palavras_chave": [
            "departamento pessoal", "folha de pagamento", "admissao",
            "rescisao", "ferias", "beneficios", "esocial", "clt",
            "recursos humanos",
        ],
    },
    "Area 4": {
        "peso": 15,
        "palavras_chave": [
            "excel", "pacote office", "word", "powerpoint", "contratos",
            "organizacao", "rotinas administrativas", "atendimento",
        ],
    },
}

# 3) PESO DA MODALIDADE REMOTA
# Bonus extra no score se a vaga for home office/remota (soma com os pesos
# acima — o total de tudo deve dar 100).
PESO_REMOTO = 5

# =============================================================================
# A partir daqui e' o motor do script — normalmente nao precisa editar nada
# abaixo desta linha.
# =============================================================================

ARQUIVO_ENTRADA = "vagas_input.xlsx"
ARQUIVO_SAIDA = "vagas_compatibilidade.xlsx"

KEYWORDS_REMOTO = ["home office", "remoto", "home-office", "trabalho remoto"]


def normalizar(texto: str) -> str:
    """Remove acentos e baixa a caixa, para comparar palavras-chave sem
    depender de acentuacao exata no texto da vaga."""
    if not isinstance(texto, str):
        return ""
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto.lower()


def contar_ocorrencias(texto_normalizado: str, termos: list) -> int:
    encontrados = 0
    for termo in termos:
        termo_norm = normalizar(termo)
        if re.search(r"\b" + re.escape(termo_norm) + r"\b", texto_normalizado):
            encontrados += 1
    return encontrados


def checar_elegibilidade(modalidade: str, local: str) -> str:
    """Verifica se a vaga e' viavel dado o tipo de contratacao e a cidade
    configuradas no bloco CONFIGURACAO."""
    modalidade_norm = normalizar(modalidade)
    local_norm = normalizar(local)

    cidades_presencial = [normalizar(c) for c in MINHA_CIDADE_PRESENCIAL]
    cidades_hibrido = [normalizar(c) for c in MINHA_CIDADE_HIBRIDO]

    eh_remoto = contar_ocorrencias(modalidade_norm, KEYWORDS_REMOTO) > 0
    eh_hibrido = "hibrido" in modalidade_norm
    eh_presencial = "presencial" in modalidade_norm

    if eh_remoto:
        return "Apta"
    if eh_hibrido:
        return "Apta" if any(c in local_norm for c in cidades_hibrido) else "Fora da area (hibrido)"
    if eh_presencial:
        return "Apta" if any(c in local_norm for c in cidades_presencial) else "Fora da area (presencial)"

    # Modalidade nao informada com clareza: nao bloqueia, mas sinaliza pra conferir
    return "Verificar modalidade"


def calcular_score(vaga: str, descricao: str, modalidade: str, local: str) -> tuple:
    texto_completo = normalizar(f"{vaga} {descricao}")
    texto_modalidade = normalizar(f"{modalidade} {local}")

    score_total = 0.0
    for grupo in GRUPOS_DE_INTERESSE.values():
        ocorrencias = contar_ocorrencias(texto_completo, grupo["palavras_chave"])
        score_total += min(ocorrencias / 4, 1) * grupo["peso"]

    if contar_ocorrencias(texto_modalidade, KEYWORDS_REMOTO) > 0:
        score_total += PESO_REMOTO

    score_total = round(score_total)

    elegibilidade = checar_elegibilidade(modalidade, local)
    if elegibilidade.startswith("Fora da area"):
        score_total = 0

    if score_total >= 70:
        classificacao = "Alta"
    elif score_total >= 40:
        classificacao = "Media"
    else:
        classificacao = "Baixa"

    return score_total, classificacao, elegibilidade


def gerar_modelo_entrada():
    """Cria um arquivo de exemplo caso o usuario ainda nao tenha um."""
    exemplo = pd.DataFrame([
        {
            "Vaga": "Titulo da vaga",
            "Empresa": "Nome da empresa",
            "Local": "Cidade, UF (ou Remoto)",
            "Modalidade": "Presencial, Hibrido ou Home office",
            "Salario": "R$ 0.000,00 (ou A combinar)",
            "Beneficios": "VR, VT, plano de saude etc.",
            "Descricao": "Cole aqui o texto de requisitos/atividades da vaga.",
            "Link": "https://link-da-vaga.com",
        }
    ])
    exemplo.to_excel(ARQUIVO_ENTRADA, index=False)
    print(f"Nao encontrei '{ARQUIVO_ENTRADA}'. Criei um modelo de exemplo — preencha com suas vagas e rode o script de novo.")


def main():
    try:
        df = pd.read_excel(ARQUIVO_ENTRADA)
    except FileNotFoundError:
        gerar_modelo_entrada()
        return

    colunas_esperadas = ["Vaga", "Empresa", "Local", "Modalidade", "Salario", "Beneficios", "Descricao", "Link"]
    for col in colunas_esperadas:
        if col not in df.columns:
            df[col] = ""

    resultados = df.apply(
        lambda row: calcular_score(row["Vaga"], row["Descricao"], row["Modalidade"], row["Local"]),
        axis=1,
    )
    df["Compatibilidade (%)"] = [r[0] for r in resultados]
    df["Classificacao"] = [r[1] for r in resultados]
    df["Elegibilidade"] = [r[2] for r in resultados]

    df = df.sort_values("Compatibilidade (%)", ascending=False).reset_index(drop=True)
    df = df[["Vaga", "Empresa", "Local", "Modalidade", "Salario", "Beneficios", "Compatibilidade (%)", "Classificacao", "Elegibilidade", "Descricao", "Link"]]

    df.to_excel(ARQUIVO_SAIDA, index=False, sheet_name="Vagas")
    formatar_planilha(ARQUIVO_SAIDA)
    print(f"Pronto! {len(df)} vaga(s) analisada(s). Resultado em '{ARQUIVO_SAIDA}'.")


def formatar_planilha(caminho: str):
    wb = load_workbook(caminho)
    ws = wb["Vagas"]

    fonte_padrao = "Arial"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name=fonte_padrao, bold=True, color="FFFFFF")

    fill_alta = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_media = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_baixa = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    class_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Classificacao":
            class_col_idx = idx

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(name=fonte_padrao)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if class_col_idx:
            valor = row[class_col_idx - 1].value
            fill = {"Alta": fill_alta, "Media": fill_media, "Baixa": fill_baixa}.get(valor)
            if fill:
                for cell in row:
                    cell.fill = fill

    larguras = {"A": 32, "B": 22, "C": 16, "D": 16, "E": 16, "F": 30, "G": 16, "H": 14, "I": 22, "J": 55, "K": 30}
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    ws.freeze_panes = "A2"
    wb.save(caminho)


if __name__ == "__main__":
    main()
